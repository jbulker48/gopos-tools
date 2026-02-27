import openpyxl
import os
from datetime import datetime

# Function to extract date from filename
def extract_date(filename):
    date_str = filename.split()[1][:10]  # Extract the date part
    return datetime.strptime(date_str, '%Y-%m-%d').strftime('%Y-%m-%d')

# Function to calculate average
def calculate_average(values):
    return sum(values) / len(values) if values else 0

# Directory containing the Excel files (same as the script's directory)
folder_path = os.path.dirname(os.path.realpath(__file__))

# Initialize counts dictionaries
dates_info = {}

# Process each file
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') and filename.startswith('kullanicilar'):
        filepath = os.path.join(folder_path, filename)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Extract date from filename
        date = extract_date(filename)

        # Initialize data for this date
        if date not in dates_info:
            dates_info[date] = {'total_rows': 0, 'avg_lisans_tutari': [], 'total_lisans_tutari_online': 0, 'lisans_tipi': {}, 'bayi': {}, 'server': {}}

        # Find the header indexes
        headers = {cell.value: cell.column - 1 for cell in ws[1]}

        # Count total rows
        total_rows = ws.max_row - 1  # Subtract 1 for the header row
        dates_info[date]['total_rows'] += total_rows

        # Calculate average and total of "Lisans Tutarı" where "Lisans Tipi" is "ONLINE"
        for row in ws.iter_rows(min_row=2):
            if row[headers['Lisans Tipi']].value == 'ONLINE':
                lisans_tutari = row[headers['Lisans Tutarı']].value
                dates_info[date]['avg_lisans_tutari'].append(lisans_tutari)
                dates_info[date]['total_lisans_tutari_online'] += lisans_tutari

        # Count "Lisans Tipi"
        for row in ws.iter_rows(min_row=2):
            lisans_tipi = row[headers['Lisans Tipi']].value
            if lisans_tipi:
                if lisans_tipi not in dates_info[date]['lisans_tipi']:
                    dates_info[date]['lisans_tipi'][lisans_tipi] = 0
                dates_info[date]['lisans_tipi'][lisans_tipi] += 1

        # Count "Bayi"
        if 'Bayi' in headers:
            for row in ws.iter_rows(min_row=2):
                bayi = row[headers['Bayi']].value
                if bayi:
                    if bayi not in dates_info[date]['bayi']:
                        dates_info[date]['bayi'][bayi] = 0
                    dates_info[date]['bayi'][bayi] += 1

        # Count "Server"
        if 'Server' in headers:
            for row in ws.iter_rows(min_row=2):
                server = row[headers['Server']].value
                if server:
                    if server not in dates_info[date]['server']:
                        dates_info[date]['server'][server] = 0
                    dates_info[date]['server'][server] += 1

# Create a new workbook for the stats
stats_wb = openpyxl.Workbook()
stats_ws = stats_wb.active
stats_ws.title = 'Stats'

# Set the headers for the stats sheet
all_lisans_tipi = set()
all_bayi = set()
all_server = set()
for date_data in dates_info.values():
    all_lisans_tipi.update(date_data['lisans_tipi'].keys())
    all_bayi.update(date_data['bayi'].keys())
    all_server.update(date_data['server'].keys())

headers = ['Date', 'Total Rows', 'Avg Lisans Tutarı (ONLINE)', 'Total Lisans Tutarı (ONLINE)'] + \
          [f'Lisans Tipi: {lt}' for lt in sorted(all_lisans_tipi)] + \
          [f'Bayi: {b}' for b in sorted(all_bayi)] + \
          [f'Server: {s}' for s in sorted(all_server)]
stats_ws.append(headers)

# Fill in the stats
for date, info in dates_info.items():
    row = [date, info['total_rows'], calculate_average(info['avg_lisans_tutari']), info['total_lisans_tutari_online']]
    row.extend([info['lisans_tipi'].get(lt, 0) for lt in sorted(all_lisans_tipi)])
    row.extend([info['bayi'].get(b, '') for b in sorted(all_bayi)])
    row.extend([info['server'].get(s, '') for s in sorted(all_server)])
    stats_ws.append(row)

# Save the stats workbook
stats_filename = f'stats_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
stats_wb.save(os.path.join(folder_path, stats_filename))
