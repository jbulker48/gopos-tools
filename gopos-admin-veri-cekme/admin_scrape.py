from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import pandas as pd
import datetime
import openpyxl


# HTML aç 
with open("page_source.html", encoding="utf-8") as fp:
    soup = BeautifulSoup(fp, 'html.parser')


# Tabloyu bul
table = soup.find("table", {"id": "customerTbl"})
columns = [i.get_text(strip=True) for i in table.find_all("th")]
kullanicilar = table.find_all("tr", class_="ecommerce-card load-more__item")

# Excel oluşturmak için liste oluştur
kullanicilar_list = []

# Her kullanıcı için döndür ve verilerini listeye ekle
for kullanici in kullanicilar:
    kullanici_dict = {}

    # Kullanıcı Adı
    isim_isletme = str(kullanici.find_all("td")[0])
    kullanici_adi = isim_isletme.split(r"<br/>")[0]
    kullanici_adi = kullanici_adi.lstrip(r"<td>")
    kullanici_adi = kullanici_adi.strip()
    kullanici_dict["Kullanıcı Adı"] = kullanici_adi

    # İşletme Adı
    isletme_adi = str(kullanici.find_all("td")[0])
    isletme_adi = isletme_adi.split(r"<br/>")[1]
    isletme_adi = isletme_adi.rstrip(r"</td>")
    isletme_adi = isletme_adi.strip()
    kullanici_dict["İşletme Adı"] = isletme_adi

    # Lisans Tutarı
    lisans_tutari = kullanici.find_all("td")[1]
    lisans_tutari = str(lisans_tutari.get_text()).strip()
    kullanici_dict["Lisans Tutarı"] = float(lisans_tutari[1:].replace('₺', '').replace('.', '').replace(',', '.'))

    # Lisans Tipi
    lisans_tipi = kullanici.find_all("td")[2]
    lisans_tipi = str(lisans_tipi.find("option", selected=True).get_text()).strip()
    kullanici_dict["Lisans Tipi"] = lisans_tipi

    # Lisans Başlangıç Tarihi
    lisans_baslangic_tarihi = kullanici.find_all("td")[3]
    lisans_baslangic_tarihi = lisans_baslangic_tarihi.find("input", {"class": "form-control", "type": "date"})
    lisans_baslangic_tarihi = lisans_baslangic_tarihi["value"]
    kullanici_dict["Lisans Başlangıç"] = lisans_baslangic_tarihi

    # Lisans Bitiş Tarihi
    lisans_bitis_tarihi = kullanici.find_all("td")[4]
    lisans_bitis_tarihi = lisans_bitis_tarihi.find("input", {"class": "form-control", "type": "date"})
    lisans_bitis_tarihi = lisans_bitis_tarihi["value"]
    kullanici_dict["Lisans Bitiş"] = lisans_bitis_tarihi

    # listeye ekle
    kullanicilar_list.append(kullanici_dict)


# DataFrame oluştur
df = pd.DataFrame(kullanicilar_list)

# Kayıt İçin Saati al, isim oluştur
now = datetime.datetime.now()
timestamp = now.strftime('%Y%m%d%H%M%S%f')[:17]
xlsx_name = f"kullanicilar {timestamp}.xlsx"

# excel dosyası olarak kaydet
df.to_excel(xlsx_name, index=False)

# Excel'i formatla
wb = load_workbook(xlsx_name)
ws = wb.active

# Yeniden boyutlandır
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# lirayı düzelt
lira_format = '₺#,##0.00'
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=ws.max_row):
    for cell in row:
        cell.number_format = lira_format

# tarihleri düzelt
date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')
# E
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, max_row=ws.max_row):
    for cell in row:
        cell.style = date_style
# F
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, max_row=ws.max_row):
    for cell in row:
        cell.style = date_style

# Lisans Kaç Günlük
new_col_index = ws.max_column + 1
ws.cell(row=1, column=new_col_index, value='Lisans Kaç Gün')
for row in range(2, ws.max_row + 1):
    date_f = ws.cell(row=row, column=6).value  # F
    date_e = ws.cell(row=row, column=5).value  # E

    if date_f and date_e:
        if isinstance(date_f, str):
            date_f = datetime.datetime.strptime(date_f, '%Y-%m-%d')
        if isinstance(date_e, str):
            date_e = datetime.datetime.strptime(date_e, '%Y-%m-%d')
        
        kalan_gun = (date_f - date_e).days
        ws.cell(row=row, column=new_col_index, value=kalan_gun)

# G1 bold
cell = ws['G1']
bold_font = Font(bold=True)
cell.font = bold_font

# Kalan Gün Ekleme
new_col_index = ws.max_column + 1
ws.cell(row=1, column=new_col_index, value='Kalan Gün')
current_date = datetime.datetime.now()
for row in range(2, ws.max_row + 1):
    date_f = ws.cell(row=row, column=6).value  # F
    if date_f:
        if isinstance(date_f, str):
            date_f = datetime.datetime.strptime(date_f, '%Y-%m-%d')
        
        remaining_days = (date_f - current_date).days
        ws.cell(row=row, column=new_col_index, value=remaining_days+1)

# H1 bold
cell = ws['H1']
bold_font = Font(bold=True)
cell.font = bold_font

# Filtre uygula
max_row = ws.max_row
max_column = ws.max_column
filter_range = f'A1:{openpyxl.utils.get_column_letter(max_column)}{max_row}'
ws.auto_filter.ref = filter_range

# Kaydet
wb.save(xlsx_name)
