import openpyxl
import os
from datetime import datetime

# Function to extract date from filename
def extract_date(filename):
    date_str = filename.split()[1][:10]  # Extract the date part
    return datetime.strptime(date_str, '%Y-%m-%d').strftime('%Y-%m-%d')

# Function to calculate average
def calculate_average(values):
    return sum(values) / len(values) if values else 0

# Directory containing the Excel files (same as the script's directory)
folder_path = os.path.dirname(os.path.realpath(__file__))

# Initialize counts dictionaries
dates_info = {}

# Process each file
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') and filename.startswith('kullanicilar'):
        filepath = os.path.join(folder_path, filename)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Extract date from filename
        date = extract_date(filename)

        # Initialize data for this date
        if date not in dates_info:
            dates_info[date] = {'total_rows': 0, 'avg_lisans_tutari': [], 'total_lisans_tutari_online': 0, 'lisans_tipi': {}, 'bayi': {}, 'server': {}}

        # Find the header indexes
        headers = {cell.value: cell.column - 1 for cell in ws[1]}

        # Count total rows
        total_rows = ws.max_row - 1  # Subtract 1 for the header row
        dates_info[date]['total_rows'] += total_rows

        # Calculate average and total of "Lisans Tutarı" where "Lisans Tipi" is "ONLINE"
        for row in ws.iter_rows(min_row=2):
            if row[headers['Lisans Tipi']].value == 'ONLINE':
                lisans_tutari = row[headers['Lisans Tutarı']].value
                dates_info[date]['avg_lisans_tutari'].append(lisans_tutari)
                dates_info[date]['total_lisans_tutari_online'] += lisans_tutari

        # Count "Lisans Tipi"
        for row in ws.iter_rows(min_row=2):
            lisans_tipi = row[headers['Lisans Tipi']].value
            if lisans_tipi:
                if lisans_tipi not in dates_info[date]['lisans_tipi']:
                    dates_info[date]['lisans_tipi'][lisans_tipi] = 0
                dates_info[date]['lisans_tipi'][lisans_tipi] += 1

        # Count "Bayi"
        if 'Bayi' in headers:
            for row in ws.iter_rows(min_row=2):
                bayi = row[headers['Bayi']].value
                if bayi:
                    if bayi not in dates_info[date]['bayi']:
                        dates_info[date]['bayi'][bayi] = 0
                    dates_info[date]['bayi'][bayi] += 1

        # Count "Server"
        if 'Server' in headers:
            for row in ws.iter_rows(min_row=2):
                server = row[headers['Server']].value
                if server:
                    if server not in dates_info[date]['server']:
                        dates_info[date]['server'][server] = 0
                    dates_info[date]['server'][server] += 1

# Create a new workbook for the stats
stats_wb = openpyxl.Workbook()
stats_ws = stats_wb.active
stats_ws.title = 'Stats'

# Set the headers for the stats sheet
all_lisans_tipi = set()
all_bayi = set()
all_server = set()
for date_data in dates_info.values():
    all_lisans_tipi.update(date_data['lisans_tipi'].keys())
    all_bayi.update(date_data['bayi'].keys())
    all_server.update(date_data['server'].keys())

headers = ['Date', 'Total Rows', 'Avg Lisans Tutarı (ONLINE)', 'Total Lisans Tutarı (ONLINE)'] + \
          [f'Lisans Tipi: {lt}' for lt in sorted(all_lisans_tipi)] + \
          [f'Bayi: {b}' for b in sorted(all_bayi)] + \
          [f'Server: {s}' for s in sorted(all_server)]
stats_ws.append(headers)

# Fill in the stats
for date, info in dates_info.items():
    row = [date, info['total_rows'], calculate_average(info['avg_lisans_tutari']), info['total_lisans_tutari_online']]
    row.extend([info['lisans_tipi'].get(lt, 0) for lt in sorted(all_lisans_tipi)])
    row.extend([info['bayi'].get(b, '') for b in sorted(all_bayi)])
    row.extend([info['server'].get(s, '') for s in sorted(all_server)])
    stats_ws.append(row)

# Save the stats workbook
stats_filename = f'stats_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
stats_wb.save(os.path.join(folder_path, stats_filename))


import openpyxl
import os
from openpyxl.chart import PieChart, LineChart, Reference, Series
from datetime import datetime

# Mevcut kodunuz burada...
# Fonksiyonlar ve veritabanı işlemleri

# Bayii'lerde "1" değeri olanları "Diğerleri" olarak toplayalım
def consolidate_bayi_data(bayi_data):
    other_count = sum(1 for count in bayi_data.values() if count == 1)
    consolidated_data = {bayi: count for bayi, count in bayi_data.items() if count > 1}
    consolidated_data['Diğerleri'] = other_count
    return consolidated_data

# Grafik eklemek için işlevler
def add_pie_chart(ws, title, start_cell):
    chart = PieChart()
    chart.title = title
    
    labels_ref = Reference(ws, min_col=start_cell[0], min_row=start_cell[1]+1, max_row=start_cell[1]+len(ws[start_cell[1]+1:ws.max_row]))
    data_ref = Reference(ws, min_col=start_cell[0]+1, min_row=start_cell[1], max_row=start_cell[1]+len(ws[start_cell[1]+1:ws.max_row]))
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    
    ws.add_chart(chart, f"{chr(start_cell[0] + 64)}{start_cell[1] + 15}")  # Chart placement 15 rows below the data

def add_line_chart(ws, title, x_label, y_label, start_cell):
    chart = LineChart()
    chart.title = title
    chart.x_axis.title = x_label
    chart.y_axis.title = y_label
    
    x_ref = Reference(ws, min_col=start_cell[0], min_row=start_cell[1]+1, max_row=start_cell[1]+len(ws[start_cell[1]+1:ws.max_row]))
    y_ref = Reference(ws, min_col=start_cell[0]+1, min_row=start_cell[1], max_row=start_cell[1]+len(ws[start_cell[1]+1:ws.max_row]))
    
    series = Series(y_ref, xvalues=x_ref, title_from_data=True)
    chart.series.append(series)
    
    ws.add_chart(chart, f"{chr(start_cell[0] + 64)}{start_cell[1] + 15}")  # Chart placement 15 rows below the data

# Yeni bir sayfa oluşturun ve istatistikleri tabloya ekleyin
new_ws = stats_wb.create_sheet(title='Charts')

# Bayi verilerini konsolide edin
consolidated_bayi_data = consolidate_bayi_data(dates_info[date]['bayi'])

# Bayi verilerini ekleyin
new_ws.append(['Bayi', 'Count'])
row = 2
for label, data in consolidated_bayi_data.items():
    new_ws.append([label, data])

# Bayi grafiği ekleyin
add_pie_chart(new_ws, 'Bayi Sayıları', (1, row))

# Server verilerini C ve D kolonlarına ekleyin
server_labels = list(dates_info[date]['server'].keys())
server_data = list(dates_info[date]['server'].values())
new_ws.append(['Server', 'Count'])
server_start_row = row + len(consolidated_bayi_data) + 2
for label, data in zip(server_labels, server_data):
    new_ws.cell(row=server_start_row, column=3).value = label
    new_ws.cell(row=server_start_row, column=4).value = data
    server_start_row += 1

# Server grafiği ekleyin
add_pie_chart(new_ws, 'Server Sayıları', (3, row + len(consolidated_bayi_data) + 2))

# ONLINE verilerini F ve G kolonlarına ekleyin
online_dates = list(dates_info.keys())
online_counts = [info['lisans_tipi'].get('ONLINE', 0) for info in dates_info.values()]
new_ws.append(['Date', 'ONLINE Count'])
online_start_row = server_start_row + 2
for date, count in zip(online_dates, online_counts):
    new_ws.cell(row=online_start_row, column=6).value = date
    new_ws.cell(row=online_start_row, column=7).value = count
    online_start_row += 1

# ONLINE grafiği ekleyin
add_line_chart(new_ws, 'ONLINE Lisanslar', 'Tarih', 'Sayı', (6, server_start_row + 2))

# Lisans tutarı verilerini H ve I kolonlarına ekleyin
total_lisans_tutari_online_values = [info['total_lisans_tutari_online'] for info in dates_info.values()]
new_ws.append(['Date', 'Total Lisans Tutarı (ONLINE)'])
license_start_row = online_start_row + 2
for date, total in zip(online_dates, total_lisans_tutari_online_values):
    new_ws.cell(row=license_start_row, column=9).value = date
    new_ws.cell(row=license_start_row, column=10).value = total
    license_start_row += 1

# Lisans tutarı grafiği ekleyin
add_line_chart(new_ws, 'Total Lisans Tutarı (ONLINE)', 'Tarih', 'Lisans Tutarı', (9, online_start_row + 2))

# Veri yazılan sütunları gizle
columns_to_hide = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
for col in columns_to_hide:
    new_ws.column_dimensions[col].hidden = True

# İstatistikler defterini kaydet
stats_wb.save(os.path.join(folder_path, stats_filename))
