from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import openpyxl
import time
import datetime
import pandas as pd

# Load Chrome Driver
cService = webdriver.ChromeService(executable_path=r'C:\Users\GoPos\Desktop\oo-python2\chromedriver.exe')  # Change this.
options = Options()
# options.add_argument(r"C:\Users\GoPos\AppData\Local\Google\Chrome\User Data")
driver = webdriver.Chrome(service=cService, options=options)

# Load Page and login
driver.get("https://admin.gopos.com.tr/Login")
username_field = driver.find_element(By.NAME, "Email")
password_field = driver.find_element(By.NAME, "Sifre")

username_field.send_keys("Admin@gopos.com")
password_field.send_keys("!AdminPos2025--@")
password_field.send_keys(Keys.RETURN)

# Load All Users Page and take html
time.sleep(2)
driver.get("https://admin.gopos.com.tr/IsletmeModul/Index?BayiID=0")
time.sleep(2)
search_field = driver.find_element(By.ID, "txt-customerSearch-New")
search_field.send_keys("a")
time.sleep(5)
search_field.send_keys(Keys.BACK_SPACE)
time.sleep(5)

page_html = driver.page_source

# Saving html as file?
now = datetime.datetime.now()
timestamp = now.strftime('%Y%m%d%H%M%S%f')[:17]
fileToWrite = open(f"page_source.html", "w", encoding="utf-8")
fileToWrite.write(page_html)
fileToWrite.close()

# Get Users
all_users_page_soup = BeautifulSoup(page_html, 'html.parser')
all_users_page_table = all_users_page_soup.find("table", {"id": "customerTbl"})
kullanicilar = all_users_page_table.find_all("tr", class_="ecommerce-card load-more__item")


# Excel oluşturmak için liste oluştur
kullanicilar_list = []

for kullanici in kullanicilar:
    kullanici_dict = {}

    # User URL
    user_url = kullanici.find_all("td")[10]
    user_url = user_url.find('a', {"class": "btn btn-sm btn-success"})
    user_url = user_url['href']
    user_url = "https://admin.gopos.com.tr" + user_url

    # Get User html
    driver.get(user_url)
    user_html = driver.page_source
    user_soup = BeautifulSoup(user_html, 'html.parser')

    # Kullanıcı Adı
    isim_isletme = str(kullanici.find_all("td")[0])
    kullanici_adi = isim_isletme.split(r"<br/>")[0]
    kullanici_adi = kullanici_adi.lstrip(r"<td>")
    kullanici_adi = kullanici_adi.strip()
    kullanici_dict["Kullanıcı Adı"] = kullanici_adi
    
    #  TODO
    print(kullanici_adi)

    # İşletme Adı
    isletme_adi = str(kullanici.find_all("td")[0])
    isletme_adi = isletme_adi.split(r"<br/>")[1]
    isletme_adi = isletme_adi.rstrip(r"</td>")
    isletme_adi = isletme_adi.strip()
    kullanici_dict["İşletme Adı"] = isletme_adi

    # User Phone
    user_col1 = user_soup.find_all("div", class_="col-lg-3")
    try:
        user_phone = user_col1[2]
        user_phone = user_phone.find('input', {"class": "form-control", "name": "anaKullanici.Telefon"})
        user_phone = user_phone["value"]
    except:
        user_phone = ""
    kullanici_dict["Telefon No"] = user_phone

    # Lisans Tutarı
    lisans_tutari = kullanici.find_all("td")[1]
    lisans_tutari = str(lisans_tutari.get_text()).strip()
    kullanici_dict["Lisans Tutarı"] = float(lisans_tutari[1:].replace('₺', '').replace('.', '').replace(',', '.'))

    # Lisans Tipi
    lisans_tipi = kullanici.find_all("td")[2]
    lisans_tipi = str(lisans_tipi.find("option", selected=True).get_text()).strip()
    kullanici_dict["Lisans Tipi"] = lisans_tipi

    # Lisans Başlangıç Tarihi
    lisans_baslangic_tarihi = kullanici.find_all("td")[3]
    lisans_baslangic_tarihi = lisans_baslangic_tarihi.find("input", {"class": "form-control", "type": "date"})
    lisans_baslangic_tarihi = lisans_baslangic_tarihi["value"]
    kullanici_dict["Lisans Başlangıç"] = lisans_baslangic_tarihi

    # Lisans Bitiş Tarihi
    lisans_bitis_tarihi = kullanici.find_all("td")[4]
    lisans_bitis_tarihi = lisans_bitis_tarihi.find("input", {"class": "form-control", "type": "date"})
    lisans_bitis_tarihi = lisans_bitis_tarihi["value"]
    kullanici_dict["Lisans Bitiş"] = lisans_bitis_tarihi

    # Lisans Kaç Günlük
    date_format = "%Y-%m-%d"
    lisans_bas = datetime.datetime.strptime(lisans_baslangic_tarihi, date_format)
    lisans_bit = datetime.datetime.strptime(lisans_bitis_tarihi, date_format)
    delta = lisans_bit - lisans_bas
    kullanici_dict["Lisans Kaç Günlük"] = delta.days

    # Lisans Kalan
    today = datetime.datetime.now().date()
    future_date = datetime.datetime.strptime(lisans_bitis_tarihi, "%Y-%m-%d").date()
    delta = future_date - today
    kullanici_dict["Lisans Kalan"] = delta.days

    # User Email
    try:
        user_email = user_col1[5]
        user_email = user_email.find('input', {"class": "form-control", "name": "temelKullanici.Email"})
        user_email = user_email["value"]
    except:
        user_email = ""
    kullanici_dict["E-Mail"] = user_email

    # User Password
    try:
        user_pass = user_col1[6]
        user_pass = user_pass.find('input', {"class": "form-control", "name": "temelKullanici.Sifre"})
        user_pass = user_pass["value"]
    except:
        user_pass = ""
    kullanici_dict["Şifre"] = user_pass

    # User Bayi
    try:
        user_bayi = user_col1[17]
        user_bayi = str(user_bayi.find("option", selected=True).get_text()).strip()
    except:
        user_bayi = ""
    kullanici_dict["Bayi"] = user_bayi

    # User Server
    try:
        user_server = user_col1[7]
        user_server = str(user_server.find("option", selected=True).get_text()).strip()
    except:
        user_server = ""
    kullanici_dict["Server"] = user_server

    # User DB Name
    try:
        user_DB_name = user_col1[8]
        user_DB_name = user_DB_name.find('input', {"class": "form-control", "name": "temelKullanici.DbName"})
        user_DB_name = user_DB_name["value"]
    except:
        user_DB_name = ""
    kullanici_dict["DB Name"] = user_DB_name

    # User DB User Name
    try:
        user_DB_user_name = user_col1[9]
        user_DB_user_name = user_DB_user_name.find('input', {"class": "form-control", "name": "temelKullanici.DbUser"})
        user_DB_user_name = user_DB_user_name["value"]
    except:
        user_DB_user_name = ""
    kullanici_dict["DB User Name"] = user_DB_user_name

    # User DB Password
    try:
        user_DB_pass = user_col1[10]
        user_DB_pass = user_DB_pass.find('input', {"class": "form-control", "name": "temelKullanici.DbPass"})
        user_DB_pass = user_DB_pass["value"]
    except:
        user_DB_pass = ""
    kullanici_dict["User DB Password"] = user_DB_pass

    # User DB Host Name
    try:
        user_DB_host_name = user_col1[11]
        user_DB_host_name = user_DB_host_name.find('input', {"class": "form-control", "name": "temelKullanici.Hostname"})
        user_DB_host_name = user_DB_host_name["value"]
    except:
        user_DB_host_name = ""
    kullanici_dict["User DB Host Name"] = user_DB_host_name

    # User URL
    kullanici_dict["URL"] = user_url

    # User Adress
    try:
        user_adress = user_col1[3]
        user_adress = user_adress.find('input', {"class": "form-control", "name": "anaKullanici.Adres"})
        user_adress = user_adress["value"]
    except:
        user_adress = ""
    kullanici_dict["Adres"] = user_adress

    # listeye ekle
    kullanicilar_list.append(kullanici_dict)
   

# DataFrame oluştur
df = pd.DataFrame(kullanicilar_list)

# Kayıt İçin Saati al, isim oluştur
now = datetime.datetime.now()
timestamp = now.strftime('%Y-%m-%d_%H-%M-%S-%f')[:23]
xlsx_name = f"kullanicilar {timestamp}.xlsx"

# excel dosyası olarak kaydet
df.to_excel(xlsx_name, index=False)

# Excel'i formatla
wb = load_workbook(xlsx_name)
ws = wb.active

# Yeniden boyutlandır
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 10)
    ws.column_dimensions[column].width = adjusted_width

# lirayı düzelt
lira_format = '₺#,##0.00'
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws.max_row):
    for cell in row:
        cell.number_format = lira_format

# Tarihleri düzelt
date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')
# F
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, max_row=ws.max_row):
    for cell in row:
        cell.style = date_style
# G
for row in ws.iter_rows(min_row=2, min_col=7, max_col=7, max_row=ws.max_row):
    for cell in row:
        cell.style = date_style

# Filtre uygula
max_row = ws.max_row
max_column = ws.max_column
filter_range = f'A1:{openpyxl.utils.get_column_letter(max_column)}{max_row}'
ws.auto_filter.ref = filter_range

# Kaydet
wb.save(xlsx_name)
